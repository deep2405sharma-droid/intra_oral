"""
validate_smartom_metadata.py
-----------------------------
Five validation checks on the SMART-OM dataset:

  CHECK 1 — Images on disk whose patient_id is NOT present in clean CSV
             (image exists, patient metadata entirely absent)

  CHECK 2 — Images on disk whose patient_id IS in clean CSV but that
             specific image file has no row in the CSV
             (patient known, but this image entry is missing)

  CHECK 3 — Rows in clean CSV whose image_path does not exist on disk
             (stale / broken path recorded in CSV)

  CHECK 4 — Columns in clean CSV where every single value is null/empty

  CHECK 5 — Field-level diff between original xlsx and clean CSV:
             for every patient present in both, report any cell whose
             value changed or was lost during cleaning

Disk structure:
    <basepath>/<target>/<folder>/<region>/<image_file>

Outputs:
    - Logger summary
    - validate.smartom.patient.report xlsx (config.ini)
"""

from src.common import intraoral_logger as iolog
from utils.load_configuration import load_config
from utils.log_handler import CustomSizeDayRotatingFileHandler

import os
import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REGION_COL_MAP = {
    "01. Dorsal tongue": "DT",
    "02. Ventral tongue": "VT",
    "03. Left buccal mucosa": "LB",
    "04. Right buccal mucosa": "RB",
    "05. Upper lip": "UL",
    "06. Lower lip": "LL",
    "07. Upper arch": "UA",
    "08. Lower arch": "LA",
    "09. Json files": "json_file",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".json"}

# Columns added by the cleaning script — not present in original xlsx
CLEANING_ADDED_COLS = {
    "label",
    "lesion_location",
    "lesion_present",
    "lesion_classification",
    "image_path",
    "source",
    "type",
}

# Rename map mirrors clean_smart_om_metadata.py exactly
COL_RENAME = {
    "S.No": "serial_no",
    "SMITA_ID": "patient_id",
    "SMITA ID": "patient_id",
    "Age": "age",
    "Sex": "sex",
    "Habit_history": "habit_history",
    "Type_of_habit": "habit_types",
    "Form_of_tobacco": "tobacco_type",
    "Smoking": "smoking_type",
    "Smoking_Type": "smoking_type",
    "Smoking_Frequency (No of times per day)": "smoking_freq_per_day",
    "Smoking_Duration of Habit (in years)": "smoking_duration_years",
    "Smoking_Habit_status": "smoking_status",
    "Smoking_Start of Habit-Age (in years)": "smoking_onset_age",
    "Chewing_Frequency (No of times per day)": "chewing_freq_per_day",
    "Chewing_Duration of Habit (in years)": "chewing_duration_years",
    "Chewing_Habit_status": "chewing_habit_status",
    "Chewing_Start of Habit-Age (in years)": "chewing_onset_age",
    "Arecanut_Frequency (No of times per day)": "arecanut_freq_per_day",
    "Arecanut_Duration of Habit (in years)": "arecanut_duration_years",
    "Arecanut_Habit_status": "arecanut_status",
    "Arecanut_Start of Habit-Age (in years)": "arecanut_onset_age",
    "Alcohol_Frequency (No of times per week)": "alcohol_freq_per_week",
    "Alcohol_Duration of Habit (in years)": "alcohol_duration_years",
    "Alcohol_Habit status": "alcohol_status",
    "Alcohol_Start of Habit-Age (in years)": "alcohol_onset_age",
    "Brushing_habit": "oral_hygiene_cleaning_aid",
    "Brushing_habit_Type of cleaning aid": "oral_hygiene_cleaning_aid",
    "Brushing_habit_Material used": "oral_hygiene_material",
    "Brushing_habit_Method of brushing": "oral_hygiene_brushing_method",
    "Brushing_habit_Frequency of brushing": "oral_hygiene_brushing_freq",
    "Brushing_habit_Duration of brushing": "oral_hygiene_brushing_duration",
    "Brushing_habit_Frequency of changing toothbrush (in months)": "oral_hygiene_brush_change_freq",
    "Family history": "family_history",
    "Past Medical history": "past_medical_history",
    "Past Dental History": "past_dental_history",
    "Denture usage": "denture_usage",
}


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def get_base_patient_id(pid: str) -> str:
    """SMITA00220_W -> SMITA00220 | SMITA00024-1 -> SMITA00024"""
    pid = str(pid).strip()
    # pid = re.sub(r"_W$", "", pid)
    # pid = re.sub(r"-\d+$", "", pid)
    return pid


def extract_patient_id_from_filename(fname: str) -> str | None:
    """Extract base patient ID from an image filename."""
    pid_match = re.match(r"^SMITA[^_-]*?(?:[_-][W1])", fname, re.IGNORECASE)
    if not pid_match:
        pid_match = re.match(r"([^_.]+)", fname)
    if pid_match:
        return pid_match.group(0)
    return None


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------


def load_clean_csv(csv_path: str) -> pd.DataFrame:
    """Load clean CSV; normalise patient_id."""
    df = pd.read_csv(csv_path, dtype=str)
    # df["patient_id"] = df["patient_id"].apply(
    #     lambda x: get_base_patient_id(x) if pd.notna(x) else x
    # )
    return df


def load_xlsx_as_clean(xlsx_path: str) -> pd.DataFrame:
    """
    Read the original xlsx and apply exactly the same transformations
    as the cleaning script (header merge, rename, null normalise, strip)
    so values are directly comparable to the clean CSV.
    Returns a DataFrame indexed by base patient_id.
    Duplicate patient rows (watchlist / repeat visits) are kept separately.
    """
    df_raw = pd.read_excel(xlsx_path, header=None)

    # Merge two-row header
    row0 = df_raw.iloc[0].tolist()
    row1 = df_raw.iloc[1].tolist()
    filled_group, last = [], None
    for v in row0:
        if pd.notna(v):
            last = str(v).strip()
        filled_group.append(last)
    headers = []
    for grp, sub in zip(filled_group, row1):
        if pd.isna(sub) or str(sub).strip() == "":
            headers.append(grp)
        else:
            sub_clean = str(sub).strip()
            headers.append(
                f"{grp}_{sub_clean}"
                if grp and not sub_clean.startswith(grp)
                else sub_clean
            )

    data = df_raw.iloc[2:].copy().reset_index(drop=True)
    data.columns = headers

    # Rename to snake_case
    data.rename(
        columns={h: COL_RENAME[h] for h in data.columns if h in COL_RENAME},
        inplace=True,
    )

    # Null normalisation
    data.replace(r"^\s*-+\s*$", np.nan, regex=True, inplace=True)
    data.replace(r"^\s*$", np.nan, regex=True, inplace=True)

    # Strip strings
    for col in data.select_dtypes(include=["object", "str"]).columns:
        data[col] = data[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # Normalise patient_id
    # data["patient_id"] = data["patient_id"].apply(
    #     lambda x: get_base_patient_id(str(x)) if pd.notna(x) else x
    # )

    return data


def collect_disk_images(
    basepath: str,
    targets: list,
    folders: list,
    regions: list,
) -> list[dict]:
    """
    Walk target/folder/region tree and return one dict per image file:
        abs_path, target, folder, region, region_col, filename, patient_id
    """
    records = []
    for target in targets:
        for folder in folders:
            for region in regions:
                dir_path = os.path.join(basepath, target, folder, region)
                if not os.path.isdir(dir_path):
                    continue
                region_col = REGION_COL_MAP.get(region)
                for fname in sorted(os.listdir(dir_path)):
                    fpath = os.path.join(dir_path, fname)
                    if not os.path.isfile(fpath):
                        continue
                    if Path(fname).suffix.lower() not in IMAGE_EXTENSIONS:
                        continue
                    records.append(
                        {
                            "abs_path": fpath,
                            "target": target,
                            "folder": folder,
                            "region": region,
                            "region_col": region_col,
                            "filename": fname,
                            "patient_id": extract_patient_id_from_filename(fname),
                        }
                    )
    return records


# ---------------------------------------------------------------------------
# CHECK 1 — Images whose patient_id is absent from clean CSV entirely
# ---------------------------------------------------------------------------


def check1_images_patient_missing_from_csv(
    disk_records: list[dict],
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    For every image on disk: is the patient_id present at all in the CSV?
    Returns records where the patient_id has zero rows in the CSV.
    """
    pids_in_csv = set(df_csv["patient_id"].dropna().unique())
    issues = []
    for rec in disk_records:
        pid = rec["patient_id"]
        if pid is None:
            issues.append(
                {**rec, "issue_detail": "Cannot parse patient ID from filename"}
            )
        elif pid not in pids_in_csv:
            issues.append({**rec, "issue_detail": f"'{pid}' has no rows in clean CSV"})
    return issues


# ---------------------------------------------------------------------------
# CHECK 2 — Images on disk that have no matching row in the clean CSV
# ---------------------------------------------------------------------------


def check2_image_entry_missing_from_csv(
    disk_records: list[dict],
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    For every image on disk: does the clean CSV contain a row whose
    image_path matches this file's absolute path?
    Only fires when the patient_id IS present in the CSV (check 1 covers the rest).
    """
    recorded_paths = set(df_csv["image_path"].dropna().str.strip())
    pids_in_csv = set(df_csv["patient_id"].dropna().unique())
    issues = []
    for rec in disk_records:
        pid = rec["patient_id"]
        if pid is None or pid not in pids_in_csv:
            continue  # already caught by check 1
        if rec["abs_path"] not in recorded_paths:
            issues.append(
                {
                    **rec,
                    "issue_detail": (
                        f"Patient '{pid}' is in CSV but this image has no row: "
                        f"{rec['abs_path']}"
                    ),
                }
            )
    return issues


# ---------------------------------------------------------------------------
# CHECK 3 — CSV rows whose image_path does not exist on disk
# ---------------------------------------------------------------------------


def check3_csv_paths_missing_on_disk(
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    For every row in the clean CSV: does the image_path file actually exist?
    Returns one record per broken / stale path.
    """
    issues = []
    for _, row in df_csv.iterrows():
        path = str(row.get("image_path", "")).strip()
        if not path or path == "nan":
            issues.append(
                {
                    "patient_id": row.get("patient_id", ""),
                    "image_path": path,
                    "issue_detail": "image_path is empty/null in CSV row",
                }
            )
        elif not os.path.isfile(path):
            issues.append(
                {
                    "patient_id": row.get("patient_id", ""),
                    "image_path": path,
                    "issue_detail": f"File does not exist on disk: {path}",
                }
            )
    return issues


# ---------------------------------------------------------------------------
# CHECK 4 — Columns in clean CSV that are entirely null/empty
# ---------------------------------------------------------------------------


def check4_fully_null_columns(df_csv: pd.DataFrame) -> list[dict]:
    """
    Returns one record per column where every value is null or empty string.
    """
    issues = []
    for col in df_csv.columns:
        series = df_csv[col].replace("", np.nan)
        if series.isna().all():
            issues.append(
                {
                    "column": col,
                    "issue_detail": f"Column '{col}' is entirely null/empty across all {len(df_csv)} rows",
                }
            )
    return issues


# ---------------------------------------------------------------------------
# CHECK 5 — Field-level diff: original xlsx vs clean CSV
# ---------------------------------------------------------------------------


# Helper function
def _safe_to_int(val):
    """Safely convert to integer, returns pd.NA on failure."""
    if pd.isna(val):
        return pd.NA
    try:
        # Handles 1.0, "1", "1.0", etc.
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return pd.NA


def check5_metadata_fidelity(
    df_xlsx: pd.DataFrame,
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    Field-level comparison between original xlsx and clean CSV.
    Handles int vs float for serial_no and age safely.
    """
    # Columns to compare (exclude columns added only during cleaning)
    compare_cols = [
        c for c in df_xlsx.columns if c not in CLEANING_ADDED_COLS and c != "patient_id"
    ]

    # Columns that must be compared numerically (serial_no, age)
    NUMERIC_COLS = {"serial_no", "age"}

    # Use first row per patient from CSV (handles duplicate visits)
    csv_meta = df_csv.drop_duplicates(subset="patient_id").set_index("patient_id")[
        compare_cols
    ]

    issues = []

    for _, xlsx_row in df_xlsx.iterrows():
        pid = xlsx_row.get("patient_id")
        if pd.isna(pid):
            continue

        pid = str(pid).strip()

        if pid not in csv_meta.index:
            issues.append(
                {
                    "patient_id": pid,
                    "column": "ALL",
                    "xlsx_value": "(row present in xlsx)",
                    "csv_value": "(absent in CSV)",
                    "issue_detail": f"'{pid}' exists in xlsx but missing in clean CSV",
                }
            )
            continue

        csv_row = csv_meta.loc[pid]

        for col in compare_cols:
            xlsx_val = xlsx_row.get(col)
            csv_val = csv_row.get(col)

            # === NUMERIC COMPARISON (serial_no, age) ===
            if col in NUMERIC_COLS:
                xlsx_num = _safe_to_int(xlsx_val)
                csv_num = _safe_to_int(csv_val)

                if pd.isna(xlsx_num) and pd.isna(csv_num):
                    continue  # both missing → OK
                if pd.isna(xlsx_num) or pd.isna(csv_num):
                    # one missing, one present → mismatch
                    issues.append(
                        {
                            "patient_id": pid,
                            "column": col,
                            "xlsx_value": xlsx_val,
                            "csv_value": csv_val,
                            "issue_detail": f"Mismatch in '{col}': xlsx={xlsx_val!r} vs csv={csv_val!r} (NA difference)",
                        }
                    )
                    continue

                if xlsx_num != csv_num:
                    issues.append(
                        {
                            "patient_id": pid,
                            "column": col,
                            "xlsx_value": xlsx_val,
                            "csv_value": csv_val,
                            "issue_detail": f"Mismatch in '{col}': xlsx={xlsx_num} vs csv={csv_num}",
                        }
                    )
                continue

            # === STRING COMPARISON (all other columns) ===
            xlsx_str = "" if pd.isna(xlsx_val) else str(xlsx_val).strip()
            csv_str = "" if pd.isna(csv_val) else str(csv_val).strip()

            if xlsx_str != csv_str:
                issues.append(
                    {
                        "patient_id": pid,
                        "column": col,
                        "xlsx_value": xlsx_val,
                        "csv_value": csv_val,
                        "issue_detail": f"Mismatch in '{col}': xlsx='{xlsx_str}' vs csv='{csv_str}'",
                    }
                )

    return issues


def check5_metadata_fidelity_old(
    df_xlsx: pd.DataFrame,
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    For every patient present in both xlsx and CSV, compare each metadata
    field value.  Reports:
      - patient in xlsx but completely absent from CSV
      - individual field mismatches (value changed or lost during cleaning)

    Comparison columns: those present in xlsx after renaming, excluding any
    column added solely by the cleaning script.

    Fix: serial_no and age are integers in the xlsx (stored as Python int)
    but loaded as float64 in the CSV (e.g. 1 vs 1.0).  Both sides are cast
    to pandas nullable Int64 before comparison so "1" == "1.0" does not
    produce a spurious mismatch.
    """
    # Columns to compare: xlsx renamed cols minus cleaning-added cols
    compare_cols = [
        c for c in df_xlsx.columns if c not in CLEANING_ADDED_COLS and c != "patient_id"
    ]

    # Columns that should be compared as integers to avoid int/float mismatches
    INT_COMPARE_COLS = {"serial_no", "age"}

    # Build per-patient unique metadata from CSV (first row per patient)
    csv_meta = (
        df_csv.drop_duplicates(subset="patient_id")[
            ["patient_id"] + compare_cols
        ].set_index("patient_id")
        if all(c in df_csv.columns for c in compare_cols)
        else df_csv.drop_duplicates(subset="patient_id")
        .set_index("patient_id")
        .reindex(columns=compare_cols)
    )

    def _to_int64(val):
        """Convert a value to pandas Int64 (nullable), or pd.NA if not numeric."""
        if pd.isna(val):
            return pd.NA
        try:
            return pd.array([val], dtype="Int64")[0]
        except (ValueError, TypeError):
            return pd.NA

    issues = []

    for _, xlsx_row in df_xlsx.iterrows():
        pid = xlsx_row.get("patient_id")
        if pd.isna(pid):
            continue

        if pid not in csv_meta.index:
            issues.append(
                {
                    "patient_id": pid,
                    "column": "ALL",
                    "xlsx_value": "(row present in xlsx)",
                    "csv_value": "(patient absent from clean CSV)",
                    "issue_detail": f"'{pid}' exists in xlsx but has no row in clean CSV",
                }
            )
            continue

        csv_row = csv_meta.loc[pid]

        for col in compare_cols:
            xlsx_val = xlsx_row.get(col)
            csv_val = csv_row.get(col) if col in csv_row.index else np.nan

            if col in INT_COMPARE_COLS:
                # Cast both sides to Int64 to avoid int vs float64 false mismatches
                xlsx_cmp = _to_int64(xlsx_val)
                csv_cmp = _to_int64(csv_val)
                # Both missing → considered equal
                if pd.isna(xlsx_cmp) and pd.isna(csv_cmp):
                    continue
                if pd.isna(xlsx_cmp) or pd.isna(csv_cmp):
                    issues.append(
                        {
                            "patient_id": pid,
                            "column": col,
                            "xlsx_value": xlsx_val,
                            "csv_value": csv_val,
                            "issue_detail": f"Mismatch in '{col}': xlsx={xlsx_val!r} vs csv={csv_val!r} (one is NA)",
                        }
                    )
                    continue
                # Both present → normal comparison
                if xlsx_cmp != csv_cmp:
                    issues.append(
                        {
                            "patient_id": pid,
                            "column": col,
                            "xlsx_value": xlsx_val,
                            "csv_value": csv_val,
                            "issue_detail": (
                                f"Mismatch in '{col}': xlsx={xlsx_cmp} vs csv={csv_cmp}"
                            ),
                        }
                    )
            else:
                # General string-normalised comparison (unchanged behaviour)
                xlsx_str = "" if pd.isna(xlsx_val) else str(xlsx_val).strip()
                csv_str = (
                    ""
                    if (pd.isna(csv_val) or str(csv_val).strip() == "nan")
                    else str(csv_val).strip()
                )

                if xlsx_str != csv_str:
                    issues.append(
                        {
                            "patient_id": pid,
                            "column": col,
                            "xlsx_value": xlsx_val,
                            "csv_value": csv_val,
                            "issue_detail": (
                                f"Mismatch in '{col}': xlsx='{xlsx_str}' vs csv='{csv_str}'"
                            ),
                        }
                    )

    return issues


# ---------------------------------------------------------------------------
# CHECK 6 — Patient ID set comparison: xlsx vs clean CSV
# ---------------------------------------------------------------------------


def check6_patient_id_set_comparison(
    df_xlsx: pd.DataFrame,
    df_csv: pd.DataFrame,
) -> dict:
    """
    Compare the unique patient_id sets between the original xlsx and the
    clean CSV.  Returns a dict with:
        common_count          : patients present in both
        only_in_csv_count     : patients in CSV but absent from xlsx
        only_in_xlsx_count    : patients in xlsx but absent from CSV
        common_ids            : sorted list of common patient_ids
        only_in_csv_ids       : sorted list of patient_ids only in CSV
        only_in_xlsx_ids      : sorted list of patient_ids only in xlsx
    """
    xlsx_ids = set(str(p).strip() for p in df_xlsx["patient_id"].dropna().unique())
    csv_ids = set(str(p).strip() for p in df_csv["patient_id"].dropna().unique())

    common = xlsx_ids & csv_ids
    only_csv = csv_ids - xlsx_ids
    only_xlsx = xlsx_ids - csv_ids

    return {
        "common_count": len(common),
        "only_in_csv_count": len(only_csv),
        "only_in_xlsx_count": len(only_xlsx),
        "common_ids": sorted(common),
        "only_in_csv_ids": sorted(only_csv),
        "only_in_xlsx_ids": sorted(only_xlsx),
    }


# ---------------------------------------------------------------------------
# Report writer
# ---------------------------------------------------------------------------


def write_excel_report(
    c1: list[dict],
    c2: list[dict],
    c3: list[dict],
    c4: list[dict],
    c5: list[dict],
    c6: dict,
    report_path: str,
) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(report_path)), exist_ok=True)

    wb = Workbook()
    HDR = Font(bold=True, color="FFFFFF")
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header(ws, hex_color: str):
        fill = PatternFill("solid", start_color=hex_color)
        for cell in ws[1]:
            cell.font = HDR
            cell.fill = fill
            cell.alignment = CTR

    def autofit(ws):
        for col_cells in ws.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
                width + 4, 70
            )

    fill_ok = PatternFill("solid", start_color="E2EFDA")
    fill_err = PatternFill("solid", start_color="FCE4D6")

    # ---- Summary ------------------------------------------------------------
    ws_s = wb.active
    ws_s.title = "Summary"
    ws_s.append(["Check", "Description", "Issue Count"])
    style_header(ws_s, "2E4057")

    summary = [
        (
            "CHECK 1",
            "Images on disk — patient_id absent from clean CSV",
            len(c1),
            False,
        ),
        (
            "CHECK 2",
            "Images on disk — image row missing from clean CSV",
            len(c2),
            False,
        ),
        ("CHECK 3", "Clean CSV rows — image_path not found on disk", len(c3), False),
        ("CHECK 4", "Clean CSV columns — entirely null/empty", len(c4), False),
        (
            "CHECK 5",
            "Metadata fidelity — field-level xlsx vs clean CSV mismatches (int fix applied)",
            len(c5),
            False,
        ),
        (
            "CHECK 6",
            (
                f"Patient ID set comparison — "
                f"Common: {c6['common_count']} | "
                f"Only in CSV: {c6['only_in_csv_count']} | "
                f"Only in XLSX: {c6['only_in_xlsx_count']}"
            ),
            c6["only_in_csv_count"] + c6["only_in_xlsx_count"],
            False,
        ),
    ]
    for check, desc, count, is_ok in summary:
        idx = ws_s.max_row + 1
        ws_s.append([check, desc, count])
        fill = fill_ok if count == 0 else fill_err
        for cell in ws_s[idx]:
            cell.fill = fill
    autofit(ws_s)

    # ---- CHECK 1 ------------------------------------------------------------
    ws1 = wb.create_sheet("C1 Patient Missing From CSV")
    IMG_COLS = [
        "patient_id",
        "filename",
        "target",
        "folder",
        "region",
        "region_col",
        "abs_path",
        "issue_detail",
    ]
    ws1.append(IMG_COLS)
    style_header(ws1, "C00000")
    for rec in c1:
        ws1.append([rec.get(c, "") for c in IMG_COLS])
    autofit(ws1)

    # ---- CHECK 2 ------------------------------------------------------------
    ws2 = wb.create_sheet("C2 Image Entry Missing")
    ws2.append(IMG_COLS)
    style_header(ws2, "ED7D31")
    for rec in c2:
        ws2.append([rec.get(c, "") for c in IMG_COLS])
    autofit(ws2)

    # ---- CHECK 3 ------------------------------------------------------------
    ws3 = wb.create_sheet("C3 Broken Image Paths")
    ws3.append(["patient_id", "image_path", "issue_detail"])
    style_header(ws3, "FF0000")
    for rec in c3:
        ws3.append(
            [
                rec.get("patient_id", ""),
                rec.get("image_path", ""),
                rec.get("issue_detail", ""),
            ]
        )
    autofit(ws3)

    # ---- CHECK 4 ------------------------------------------------------------
    ws4 = wb.create_sheet("C4 Null Columns")
    ws4.append(["column", "issue_detail"])
    style_header(ws4, "7030A0")
    for rec in c4:
        ws4.append([rec.get("column", ""), rec.get("issue_detail", "")])
    autofit(ws4)

    # ---- CHECK 5 ------------------------------------------------------------
    ws5 = wb.create_sheet("C5 Metadata Fidelity")
    ws5.append(["patient_id", "column", "xlsx_value", "csv_value", "issue_detail"])
    style_header(ws5, "0070C0")
    for rec in c5:
        ws5.append(
            [
                rec.get("patient_id", ""),
                rec.get("column", ""),
                rec.get("xlsx_value", ""),
                rec.get("csv_value", ""),
                rec.get("issue_detail", ""),
            ]
        )
    autofit(ws5)

    # ---- CHECK 6 — Patient ID set comparison --------------------------------
    ws6 = wb.create_sheet("C6 Patient ID Set Comparison")
    ws6.append(["Category", "Count", "Patient IDs"])
    style_header(ws6, "375623")

    def _ids_cell(ids: list) -> str:
        return ", ".join(ids)

    ws6.append(
        [
            "Common patient_ids (in both xlsx and CSV)",
            c6["common_count"],
            _ids_cell(c6["common_ids"]),
        ]
    )
    ws6.append(
        [
            "Only in CSV (present in clean CSV, missing from xlsx)",
            c6["only_in_csv_count"],
            _ids_cell(c6["only_in_csv_ids"]),
        ]
    )
    ws6.append(
        [
            "Only in XLSX (present in xlsx, missing from clean CSV)",
            c6["only_in_xlsx_count"],
            _ids_cell(c6["only_in_xlsx_ids"]),
        ]
    )

    # Colour rows
    green = PatternFill("solid", start_color="E2EFDA")
    orange = PatternFill("solid", start_color="FCE4D6")
    yellow = PatternFill("solid", start_color="FFEB9C")
    row_fills = [green, orange, yellow]
    for row_idx, fill in zip(range(2, 5), row_fills):
        for cell in ws6[row_idx]:
            cell.fill = fill

    # Auto-wrap the IDs column
    ws6.column_dimensions["A"].width = 55
    ws6.column_dimensions["B"].width = 12
    ws6.column_dimensions["C"].width = 80
    for row in ws6.iter_rows(min_row=2, max_row=4, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Also add individual-row detail sheets for easy review
    for sheet_name, key, color in [
        ("C6a Common IDs", "common_ids", "375623"),
        ("C6b Only In CSV IDs", "only_in_csv_ids", "ED7D31"),
        ("C6c Only In XLSX IDs", "only_in_xlsx_ids", "C00000"),
    ]:
        ws_sub = wb.create_sheet(sheet_name)
        ws_sub.append(["#", "patient_id"])
        style_header(ws_sub, color)
        for i, pid in enumerate(c6[key], start=1):
            ws_sub.append([i, pid])
        autofit(ws_sub)

    wb.save(report_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def initialize_logger(config):
    log_filename = config.get("LOGGER", "logger.filename")
    logger = iolog.getLogger(log_filename)
    return logger


if __name__ == "__main__":
    config = load_config()
    logger = initialize_logger(config=config)

    basepath = config.get("SMART-OM-DATAPATH", "smartom.basepath")
    xlsx_path = os.path.join(
        basepath, config.get("SMART-OM-DATAPATH", "smartom.metadata.file")
    )
    clean_csv = config.get(
        "SMART-OM-DATAPATH", "smartom.patient.clean.metadata.filename"
    )
    report_path = config.get("VALIDATE", "validate.smartom.patient.report")

    targets = [
        config.get("SMART-OM-DATAPATH", "smartom.normal"),
        config.get("SMART-OM-DATAPATH", "smartom.variation"),
        config.get("SMART-OM-DATAPATH", "smartom.opmd"),
    ]
    folders = [
        f.strip() for f in config.get("SMART-OM-DATAPATH", "smartom.folders").split(",")
    ]
    regions = [
        r.strip() for r in config.get("SMART-OM-DATAPATH", "smartom.regions").split(",")
    ]

    logger.info("=== SMART-OM Metadata Validation ===")
    logger.info(f"Basepath  : {basepath}")
    logger.info(f"XLSX      : {xlsx_path}")
    logger.info(f"Clean CSV : {clean_csv}")

    logger.info("Loading clean CSV...")
    df_csv = load_clean_csv(clean_csv)
    logger.info(f"  {len(df_csv)} rows, {len(df_csv.columns)} columns")

    logger.info("Loading and normalising original xlsx...")
    df_xlsx = load_xlsx_as_clean(xlsx_path)
    logger.info(f"  {len(df_xlsx)} patient rows")

    logger.info("Scanning disk for images...")
    disk_records = collect_disk_images(basepath, targets, folders, regions)
    logger.info(f"  {len(disk_records)} image/json files found")

    logger.info("CHECK 1 — patient_id missing from clean CSV...")
    c1 = check1_images_patient_missing_from_csv(disk_records, df_csv)
    logger.info(f"  Issues: {len(c1)}")

    logger.info("CHECK 2 — image row missing from clean CSV...")
    c2 = check2_image_entry_missing_from_csv(disk_records, df_csv)
    logger.info(f"  Issues: {len(c2)}")

    logger.info("CHECK 3 — CSV image_path not found on disk...")
    c3 = check3_csv_paths_missing_on_disk(df_csv)
    logger.info(f"  Issues: {len(c3)}")

    logger.info("CHECK 4 — fully null columns in clean CSV...")
    c4 = check4_fully_null_columns(df_csv)
    logger.info(f"  Issues: {len(c4)}")
    if c4:
        for rec in c4:
            logger.info(f"  Null column: {rec['column']}")

    logger.info("CHECK 5 — field-level metadata fidelity (xlsx vs CSV)...")
    c5 = check5_metadata_fidelity(df_xlsx, df_csv)
    logger.info(f"  Issues: {len(c5)}")

    logger.info("CHECK 6 — patient ID set comparison (xlsx vs CSV)...")
    c6 = check6_patient_id_set_comparison(df_xlsx, df_csv)
    logger.info(f"  Common patient IDs          : {c6['common_count']}")
    logger.info(f"  Only in CSV (missing xlsx)  : {c6['only_in_csv_count']}")
    logger.info(f"  Only in XLSX (missing CSV)  : {c6['only_in_xlsx_count']}")

    logger.info("Writing report...")
    write_excel_report(c1, c2, c3, c4, c5, c6, report_path)
    logger.info(f"Report saved -> {report_path}")
    logger.info("Validation complete.")
