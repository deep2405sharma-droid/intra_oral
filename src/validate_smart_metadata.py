"""
validate_cleaning.py
────────────────────
Validates the SMART-II cleaned CSV against the original XLSX.

The cleaned CSV now has ONE ROW PER IMAGE rather than one row per patient,
so validation logic is adapted accordingly:

  CHECK 1 — Images on disk whose patient_id is absent from clean CSV
  CHECK 2 — Images on disk that have no matching image_path row in clean CSV
  CHECK 3 — Clean CSV rows whose image_path does not exist on disk
  CHECK 4 — Columns in clean CSV that are entirely null/empty
  CHECK 5 — Field-level metadata fidelity: xlsx patient values vs CSV
             (deduplicates CSV to one row per patient before comparing)

Usage:
    python validate_cleaning.py
    python validate_cleaning.py --xlsx <path> --csv <path> --output <path>
"""

import argparse
import glob
import os
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.load_configuration import load_config


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN MAPPING  (xlsx col → csv col)
# ══════════════════════════════════════════════════════════════════════════════
COLUMN_MAP: dict = {
    "SMITA ID": "patient_id",
    "Age in years": "age",
    "Sex": "sex",
    "Religion": "religion",
    "Marital status": "marital_status",
    "Education": "education",
    "Occupation": "occupation",
    "Diet history": "diet",
    "Spice consumption": "spice_consumption",
    "Habit history": "habit_history",
    "If yes, mention habits": "habit_types",
    "If Tobacco, mention the type": "tobacco_type",
    "SMOKING FORM OF TOBACCO | Type": "smoking_type",
    "SMOKING FORM OF TOBACCO | Frequency (Number of times per day)": "smoking_freq_per_day",
    "SMOKING FORM OF TOBACCO | Duration (in years)": "smoking_duration_years",
    "SMOKING FORM OF TOBACCO | Habit status": "smoking_status",
    "SMOKING FORM OF TOBACCO | Age of onset (in years)": "smoking_onset_age",
    "SMOKELESS FORM OF TOBACCO | Frequency (Number of times per day)": "chewing_freq_per_day",
    "SMOKELESS FORM OF TOBACCO | Duration (in years)": "chewing_duration_years",
    "SMOKELESS FORM OF TOBACCO | Habit status": "chewing_habit_status",
    "SMOKELESS FORM OF TOBACCO | Age of onset (in years)": "chewing_onset_age",
    "ARECANUT | Type": "arecanut_type",
    "ARECANUT | Frequency (Number of times per day)": "arecanut_freq_per_day",
    "ARECANUT | Duration (in years)": "arecanut_duration_years",
    "ARECANUT | Habit status": "arecanut_status",
    "ARECANUT | Age of onset (in years)": "arecanut_onset_age",
    "ORAL HYGIENE PRACTICES | Type of cleaning aid": "oral_hygiene_cleaning_aid",
    "ORAL HYGIENE PRACTICES | Material used": "oral_hygiene_material",
    "ORAL HYGIENE PRACTICES | Method of brushing": "oral_hygiene_brushing_method",
    "ORAL HYGIENE PRACTICES | Frequency of brushing": "oral_hygiene_brushing_freq",
    "ORAL HYGIENE PRACTICES | Duration of brushing": "oral_hygiene_brushing_duration",
    "ORAL HYGIENE PRACTICES | Frequency of changing Tooth brush": "oral_hygiene_brush_change_freq",
    "ORAL HYGIENE PRACTICES | Usage of any other Oral hygiene aids": "oral_hygiene_other_aids",
    "Family history": "family_history",
    "Past Medical history": "past_medical_history",
    "Past Dental History": "past_dental_history",
    "DENTURE HISTORY | Denture usage": "denture_usage",
    "Presence or Absence of lesion": "lesion_present",
    "Location": "lesion_location",
    "Colour": "lesion_colour",
    "Margin": "lesion_margin",
    "Surface feature": "lesion_surface_feature",
    "Description of the Lesion": "lesion_description",
    "Lesion classification": "lesion_classification",
    "Size in cms (Length x Width)": "lesion_size_cm",
    "Associated with pain": "lesion_pain",
    "Others": "other_findings",
}

# XLSX columns intentionally dropped during cleaning
XLSX_DROPPED: list = [
    "S.No",
    "SMOKING FORM OF TOBACCO | Age of quitting (in years)",
    "SMOKELESS FORM OF TOBACCO | Age of quitting (in years)",  # chewing_quit_age kept in CSV
    "ARECANUT | Name of the product",
    "ARECANUT | Quantity (Number of sachets used per day)",
    "ARECANUT | Age of quitting (in years)",
    "ARECANUT | Others",
    "ALCOHOL | Type",
    "ALCOHOL | Frequency (Number of times per week)",
    "ALCOHOL | Duration (in years)",
    "ALCOHOL | Quantity (mL/week)",
    "ALCOHOL | Habit status",
    "ALCOHOL | Age of onset (in years)",
    "ALCOHOL | Age of quitting (in years)",
    "ORAL HYGIENE PRACTICES | If Yes, mention",
    "DENTURE HISTORY | Duration of Denture usage (in years)",
    "DENTURE HISTORY | Type of Denture",
    "DENTURE HISTORY | Material used",
    "DENTURE HISTORY | Does it hurts?",
    "If multiple lesion present, mention number of sites involved",
    "If yes, duration of pain (in days)",
    "Type of pain",
]

# CSV-only columns added during cleaning — not compared against xlsx
# Includes columns retained from xlsx that were previously in DROP_COLS
# but are now kept because dynamic null-dropping is used instead.
CSV_EXTRA_COLS: list = [
    "label",
    "image_path",
    "source",
    "type",
    # intermediate column; correct boolean lesion_present is derived from label
    "lesion_present_raw",
    # retained near-empty cols (kept by dynamic null-drop, previously in DROP_COLS)
    "serial_no",
    "chewing_quit_age",
    "arecanut_product_name",
    "arecanut_quantity_sachets_per_day",
    "arecanut_other",
    "oral_hygiene_other_aids_details",
    "denture_duration_years",
    "denture_type",
    "denture_material",
    "denture_hurts",
    "lesion_num_sites",
    "lesion_pain_duration_days",
    "lesion_pain_type",
]

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".json"}


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════


def _flatten_col(col: tuple) -> str:
    top, sub = col
    return top.strip() if "Unnamed" in str(sub) else f"{top.strip()} | {sub.strip()}"


def load_xlsx(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=[0, 1])
    df.columns = [_flatten_col(c) for c in df.columns]
    return df


def load_csv(path: str) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str)


def normalize(val) -> str | None:
    """Normalise a cell value to lowercase string or None for comparison."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip().lower()
    return None if s in ("nan", "none", "nat", "", "-", "\u2013") else s


# ══════════════════════════════════════════════════════════════════════════════
# CHECK 1 — Images on disk whose patient_id is absent from clean CSV
# ══════════════════════════════════════════════════════════════════════════════


def check1_images_patient_missing_from_csv(
    images_dir: str,
    df_csv: pd.DataFrame,
    labels: list,
) -> list[dict]:
    """
    Scan every Unannotated folder on disk. For each image, check whether
    the patient_id has ANY row in the clean CSV.
    """
    pids_in_csv = set(df_csv["patient_id"].dropna().astype(str).str.strip())
    issues = []

    for label in labels:
        label_root = Path(images_dir) / label
        if not label_root.exists():
            continue
        for pid_dir in sorted(label_root.iterdir()):
            if not pid_dir.is_dir():
                continue
            pid = pid_dir.name
            unannotated = pid_dir / "Unannotated"
            if not unannotated.exists():
                continue
            for fpath in sorted(unannotated.iterdir()):
                if not fpath.is_file():
                    continue
                if fpath.suffix.lower() not in IMAGE_EXTENSIONS:
                    continue
                if pid not in pids_in_csv:
                    issues.append(
                        {
                            "patient_id": pid,
                            "label": label,
                            "filename": fpath.name,
                            "abs_path": str(fpath),
                            "issue_detail": f"'{pid}' has no rows in clean CSV",
                        }
                    )
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# CHECK 2 — Images on disk with no matching image_path row in clean CSV
# ══════════════════════════════════════════════════════════════════════════════


def check2_image_entry_missing_from_csv(
    images_dir: str,
    df_csv: pd.DataFrame,
    labels: list,
) -> list[dict]:
    """
    For images whose patient IS in the CSV, verify a row exists whose
    image_path matches the exact file path on disk.
    """
    pids_in_csv = set(df_csv["patient_id"].dropna().astype(str).str.strip())
    recorded_paths = set(df_csv["image_path"].dropna().astype(str).str.strip())
    issues = []
    for label in labels:
        label_root = Path(images_dir) / label
        if not label_root.exists():
            continue
        for pid_dir in sorted(label_root.iterdir()):
            if not pid_dir.is_dir():
                continue
            pid = pid_dir.name
            if pid not in pids_in_csv:
                continue  # already caught by check 1
            unannotated = pid_dir / "Unannotated"
            if not unannotated.exists():
                continue
            for fpath in sorted(unannotated.iterdir()):
                if not fpath.is_file():
                    continue
                if fpath.suffix.lower() not in IMAGE_EXTENSIONS:
                    continue
                if str(fpath) not in recorded_paths:
                    issues.append(
                        {
                            "patient_id": pid,
                            "label": label,
                            "filename": fpath.name,
                            "abs_path": str(fpath),
                            "issue_detail": f"Patient '{pid}' in CSV but this image has no row",
                        }
                    )
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# CHECK 3 — CSV rows whose image_path does not exist on disk
# ══════════════════════════════════════════════════════════════════════════════


def check3_csv_paths_missing_on_disk(df_csv: pd.DataFrame) -> list[dict]:
    issues = []
    for _, row in df_csv.iterrows():
        path = str(row.get("image_path", "")).strip()
        if not path or path == "nan":
            issues.append(
                {
                    "patient_id": row.get("patient_id", ""),
                    "image_path": path,
                    "issue_detail": "image_path is empty/null in CSV row",
                }
            )
        elif not os.path.isfile(path):
            issues.append(
                {
                    "patient_id": row.get("patient_id", ""),
                    "image_path": path,
                    "issue_detail": f"File does not exist on disk: {path}",
                }
            )
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# CHECK 4 — Columns in clean CSV that are entirely null/empty
# ══════════════════════════════════════════════════════════════════════════════


def check4_fully_null_columns(df_csv: pd.DataFrame) -> list[dict]:
    issues = []
    for col in df_csv.columns:
        series = df_csv[col].replace("", np.nan)
        if series.isna().all():
            issues.append(
                {
                    "column": col,
                    "issue_detail": (
                        f"Column '{col}' is entirely null/empty "
                        f"across all {len(df_csv)} rows"
                    ),
                }
            )
    return issues


# ══════════════════════════════════════════════════════════════════════════════
# CHECK 5 — Field-level metadata fidelity: xlsx vs CSV (patient-level)
# ══════════════════════════════════════════════════════════════════════════════


def check5_metadata_fidelity(
    xlsx: pd.DataFrame,
    df_csv: pd.DataFrame,
) -> list[dict]:
    """
    Deduplicate the CSV to one row per patient_id (takes first occurrence),
    then compare every mapped field value against the xlsx row.

    Note: lesion_present in CSV is a boolean derived from the label; the xlsx
    has 'Present'/'Absent' text.  We treat this as a known transformation
    and report it as TRANSFORMED rather than LOSS.
    """
    # Deduplicate CSV to one row per patient
    csv_dedup = df_csv.drop_duplicates(subset="patient_id").set_index("patient_id")

    xlsx_idx = xlsx.copy()
    xlsx_idx["SMITA ID"] = xlsx_idx["SMITA ID"].astype(str).str.strip()
    xlsx_idx = xlsx_idx.set_index("SMITA ID")

    common_pids = xlsx_idx.index.intersection(csv_dedup.index)
    issues = []

    # Patients in xlsx but completely absent from CSV
    for pid in xlsx_idx.index:
        if pid not in csv_dedup.index:
            issues.append(
                {
                    "patient_id": pid,
                    "column": "ALL",
                    "xlsx_value": "(row present in xlsx)",
                    "csv_value": "(patient absent from clean CSV)",
                    "issue_detail": f"'{pid}' exists in xlsx but has no row in clean CSV",
                }
            )

    # Field-level comparison for patients present in both
    for pid in common_pids:
        for xlsx_col, csv_col in COLUMN_MAP.items():
            if xlsx_col == "SMITA ID" or csv_col not in csv_dedup.columns:
                continue
            if xlsx_col not in xlsx_idx.columns:
                continue

            xval = normalize(xlsx_idx.at[pid, xlsx_col])
            cval = normalize(csv_dedup.at[pid, csv_col])

            if xval == cval:
                continue
            if xval is not None and cval is None:
                issues.append(
                    {
                        "patient_id": pid,
                        "column": csv_col,
                        "xlsx_value": xlsx_idx.at[pid, xlsx_col],
                        "csv_value": csv_dedup.at[pid, csv_col],
                        "issue_detail": f"DATA LOSS in '{csv_col}': xlsx='{xval}' lost in CSV",
                    }
                )
            elif xval is not None and cval is not None and xval != cval:
                issues.append(
                    {
                        "patient_id": pid,
                        "column": csv_col,
                        "xlsx_value": xlsx_idx.at[pid, xlsx_col],
                        "csv_value": csv_dedup.at[pid, csv_col],
                        "issue_detail": f"TRANSFORMED '{csv_col}': xlsx='{xval}' -> csv='{cval}'",
                    }
                )

    return issues


# ══════════════════════════════════════════════════════════════════════════════
# EXISTING CHECKS (preserved from original)
# ══════════════════════════════════════════════════════════════════════════════


def check_column_coverage(xlsx: pd.DataFrame, csv: pd.DataFrame) -> dict:
    mapped_xlsx = set(COLUMN_MAP.keys())
    all_xlsx = set(xlsx.columns)
    all_csv = set(csv.columns)
    return {
        "unmapped_xlsx": sorted(all_xlsx - mapped_xlsx - set(XLSX_DROPPED)),
        "unmapped_csv": sorted(
            all_csv - set(COLUMN_MAP.values()) - set(CSV_EXTRA_COLS)
        ),
        "missing_in_csv": {
            xc: cc for xc, cc in COLUMN_MAP.items() if cc not in all_csv
        },
    }


def check_dropped_columns(xlsx: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col in XLSX_DROPPED:
        if col not in xlsx.columns:
            rows.append(
                {
                    "xlsx_column": col,
                    "non_null_count": "N/A",
                    "verdict": "COL_NOT_FOUND",
                }
            )
            continue
        non_null = xlsx[col].dropna()
        non_null = non_null[
            ~non_null.astype(str)
            .str.strip()
            .str.lower()
            .isin(["nan", "none", "-", "\u2013", ""])
        ]
        rows.append(
            {
                "xlsx_column": col,
                "non_null_count": len(non_null),
                "verdict": "HAD_DATA" if len(non_null) > 0 else "WAS_EMPTY",
            }
        )
    return pd.DataFrame(rows)


def generate_missing_report(csv: pd.DataFrame) -> pd.DataFrame:
    total = len(csv)
    rows = [
        {
            "column": col,
            "missing_n": int(csv[col].isna().sum()),
            "missing_pct": (
                round(csv[col].isna().sum() / total * 100, 1) if total else 0.0
            ),
        }
        for col in csv.columns
    ]
    return (
        pd.DataFrame(rows)
        .sort_values("missing_pct", ascending=False)
        .reset_index(drop=True)
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT STYLES
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
LOSS_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
TRANS_FILL = PatternFill("solid", fgColor="DDEBF7")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
BODY_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(bold=True, name="Arial", size=9)
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style(cell, fill=None, bold=False, center=False):
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = Alignment(
        horizontal="center" if center else "left", vertical="center", wrap_text=True
    )
    cell.border = BORDER
    if fill:
        cell.fill = fill


def _hdr_row(ws, row_num: int, headers: list):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER


def _widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, title: str, subtitle: str = None, row: int = 1):
    ws.row_dimensions[row].height = 28
    tc = ws.cell(row=row, column=1, value=title)
    tc.font = Font(bold=True, name="Arial", size=14, color="1F4E79")
    tc.alignment = Alignment(vertical="center")
    if subtitle:
        ws.row_dimensions[row + 1].height = 18
        sc = ws.cell(row=row + 1, column=1, value=subtitle)
        sc.font = Font(name="Arial", size=9, italic=True, color="595959")


# ══════════════════════════════════════════════════════════════════════════════
# REPORT SHEETS
# ══════════════════════════════════════════════════════════════════════════════


def _sheet_summary(wb, row_check, col_coverage, c1, c2, c3, c4, c5, dropped_df):
    ws = wb.create_sheet("Summary")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    _title(
        ws,
        "Data Cleaning Validation Report  (SMART-II)",
        f"Generated {now}  |  XLSX vs cleaned CSV  |  one-row-per-image structure",
    )

    d_had = (dropped_df["verdict"] == "HAD_DATA").sum()
    overall = (
        len(c1) == 0
        and len(c2) == 0
        and len(c3) == 0
        and len(c4) == 0
        and len(c5) == 0
        and d_had == 0
    )
    verdict = (
        "CLEANING VERIFIED - No issues detected"
        if overall
        else "ISSUES FOUND - Review highlighted sheets"
    )

    ws.row_dimensions[3].height = 30
    vc = ws.cell(row=3, column=1, value=verdict)
    vc.font = Font(
        bold=True, name="Arial", size=12, color="375623" if overall else "9C0006"
    )
    vc.fill = OK_FILL if overall else LOSS_FILL
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = BORDER
    ws.merge_cells("A3:B3")

    metrics = [
        ("METRIC", "VALUE"),
        ("XLSX original patient rows", row_check["xlsx_rows"]),
        ("CSV total rows (one per image)", row_check["csv_rows"]),
        ("", ""),
        ("CHECK 1 — Images: patient_id absent from CSV", len(c1)),
        ("CHECK 2 — Images: row missing from CSV", len(c2)),
        ("CHECK 3 — CSV rows: image_path not on disk", len(c3)),
        ("CHECK 4 — CSV columns: entirely null", len(c4)),
        ("CHECK 5 — Metadata fidelity mismatches", len(c5)),
        ("", ""),
        ("Dropped XLSX columns total", len(XLSX_DROPPED)),
        ("Dropped columns that HAD data", int(d_had)),
        ("", ""),
        ("Unmapped XLSX cols", len(col_coverage["unmapped_xlsx"])),
        ("Unmapped CSV cols", len(col_coverage["unmapped_csv"])),
    ]
    for i, (label, val) in enumerate(metrics):
        r = 5 + i
        lc = ws.cell(row=r, column=1, value=label)
        vc2 = ws.cell(row=r, column=2, value=val)
        if label == "METRIC":
            _style(lc, HDR_FILL, bold=True)
            lc.font = HDR_FONT
            _style(vc2, HDR_FILL, bold=True)
            vc2.font = HDR_FONT
        elif label:
            is_err = isinstance(val, int) and val > 0 and label.startswith("CHECK")
            fill = LOSS_FILL if is_err else (ALT_FILL if i % 2 == 0 else None)
            _style(lc, fill, bold=True)
            _style(vc2, fill)

    _widths(ws, [55, 22])
    return overall


def _sheet_image_check(wb, records: list, sheet_title: str, color: str, cols: list):
    ws = wb.create_sheet(sheet_title)
    _title(ws, sheet_title)
    _hdr_row(ws, 3, cols)
    for i, rec in enumerate(records):
        r = 4 + i
        fill = PatternFill("solid", start_color=color)
        for c, key in enumerate(cols, 1):
            _style(ws.cell(row=r, column=c, value=rec.get(key, "")), fill)
    _widths(ws, [18, 14, 18, 60, 70])


def _sheet_fidelity(wb, c5: list):
    ws = wb.create_sheet("C5 Metadata Fidelity")
    _title(
        ws,
        "Field-Level Metadata Fidelity",
        "Compares xlsx patient values against deduplicated CSV rows",
    )
    cols = ["patient_id", "column", "xlsx_value", "csv_value", "issue_detail"]
    _hdr_row(ws, 3, cols)
    for i, rec in enumerate(c5):
        r = 4 + i
        fill = LOSS_FILL if "DATA LOSS" in rec.get("issue_detail", "") else TRANS_FILL
        for c, key in enumerate(cols, 1):
            _style(ws.cell(row=r, column=c, value=rec.get(key, "")), fill)
    _widths(ws, [18, 30, 25, 25, 65])


def _sheet_null_cols(wb, c4: list):
    ws = wb.create_sheet("C4 Null Columns")
    _title(ws, "Entirely Null / Empty Columns in Clean CSV")
    _hdr_row(ws, 3, ["column", "issue_detail"])
    for i, rec in enumerate(c4):
        r = 4 + i
        for c, key in enumerate(["column", "issue_detail"], 1):
            _style(ws.cell(row=r, column=c, value=rec.get(key, "")), WARN_FILL)
    _widths(ws, [30, 65])


def _sheet_dropped(wb, dropped_df: pd.DataFrame):
    ws = wb.create_sheet("Dropped XLSX Columns")
    _title(
        ws,
        "XLSX Columns Dropped During Cleaning",
        "HAD_DATA = xlsx had values — confirm drop was intentional",
    )
    _hdr_row(ws, 3, ["XLSX Column", "Non-Null Values in XLSX", "Verdict"])
    for i, row in dropped_df.iterrows():
        fill = LOSS_FILL if row["verdict"] == "HAD_DATA" else OK_FILL
        for c, v in enumerate(
            [row["xlsx_column"], row["non_null_count"], row["verdict"]], 1
        ):
            _style(ws.cell(row=4 + i, column=c, value=v), fill)
    _widths(ws, [55, 24, 18])


def _sheet_missing(wb, missing_df: pd.DataFrame):
    ws = wb.create_sheet("CSV Missing Values")
    _title(
        ws,
        "CSV Column Missing Value Report",
        "High missing % is expected for conditional sub-fields (habits, lesion details)",
    )
    _hdr_row(ws, 3, ["CSV Column", "Missing Count", "Missing %", "Assessment"])
    for i, row in missing_df.iterrows():
        pct = row["missing_pct"]
        assessment = (
            "Fully empty - investigate"
            if pct == 100
            else (
                "Sub-field (conditional)"
                if pct > 80
                else "Partial data" if pct > 0 else "Complete"
            )
        )
        fill = (
            LOSS_FILL
            if pct == 100
            else WARN_FILL if pct > 50 else TRANS_FILL if pct > 0 else OK_FILL
        )
        for c, v in enumerate(
            [row["column"], row["missing_n"], f"{pct}%", assessment], 1
        ):
            _style(ws.cell(row=4 + i, column=c, value=v), fill)
    _widths(ws, [35, 16, 14, 28])


def _sheet_unmapped(wb, col_coverage: dict):
    ws = wb.create_sheet("Unmapped Columns")
    _title(
        ws,
        "Columns Without Explicit Mapping",
        "New XLSX columns may need to be added to COLUMN_MAP or XLSX_DROPPED",
    )
    r = 3
    for heading, cols, fill in [
        (
            "XLSX columns NOT in COLUMN_MAP and NOT in XLSX_DROPPED",
            col_coverage["unmapped_xlsx"],
            WARN_FILL,
        ),
        (
            "CSV columns NOT in COLUMN_MAP and NOT in CSV_EXTRA_COLS",
            col_coverage["unmapped_csv"],
            TRANS_FILL,
        ),
    ]:
        ws.cell(row=r, column=1, value=heading).font = BOLD_FONT
        r += 1
        _hdr_row(ws, r, ["Column Name"])
        r += 1
        for col in cols or ["— none —"]:
            _style(ws.cell(row=r, column=1, value=col), fill if cols else OK_FILL)
            r += 1
        r += 1
    _widths(ws, [75])


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════


def run(xlsx_path: str, csv_path: str, images_dir: str, output_path: str) -> bool:
    sep = "=" * 60
    print(f"\n{sep}")
    print("  Patient Data Cleaning Validator  (SMART-II)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(sep)

    print("Loading files...")
    xlsx = load_xlsx(xlsx_path)
    csv = load_csv(csv_path)
    print(f"  XLSX: {xlsx.shape[0]} patient rows × {xlsx.shape[1]} cols")
    print(f"  CSV : {csv.shape[0]} rows × {csv.shape[1]} cols  (one-per-image)")

    labels = ["Normal", "OPMD", "Variation", "Unknown"]

    print("\nRunning checks...")
    c1 = check1_images_patient_missing_from_csv(images_dir, csv, labels)
    print(f"  CHECK 1 — patient_id absent from CSV          : {len(c1)}")

    c2 = check2_image_entry_missing_from_csv(images_dir, csv, labels)
    print(f"  CHECK 2 — image row missing from CSV          : {len(c2)}")

    c3 = check3_csv_paths_missing_on_disk(csv)
    print(f"  CHECK 3 — CSV image_path not on disk          : {len(c3)}")

    c4 = check4_fully_null_columns(csv)
    print(f"  CHECK 4 — entirely null columns               : {len(c4)}")
    if c4:
        for rec in c4:
            print(f"    Null col: {rec['column']}")

    c5 = check5_metadata_fidelity(xlsx, csv)
    print(f"  CHECK 5 — metadata fidelity mismatches        : {len(c5)}")

    col_coverage = check_column_coverage(xlsx, csv)
    dropped_df = check_dropped_columns(xlsx)
    missing_df = generate_missing_report(csv)

    row_check = {"xlsx_rows": len(xlsx), "csv_rows": len(csv)}

    print("\nBuilding Excel report...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    overall = _sheet_summary(
        wb, row_check, col_coverage, c1, c2, c3, c4, c5, dropped_df
    )

    _sheet_image_check(
        wb,
        c1,
        "C1 Patient Missing From CSV",
        "FFC7CE",
        ["patient_id", "label", "filename", "abs_path", "issue_detail"],
    )
    _sheet_image_check(
        wb,
        c2,
        "C2 Image Row Missing",
        "FFEB9C",
        ["patient_id", "label", "filename", "abs_path", "issue_detail"],
    )
    _sheet_image_check(
        wb,
        c3,
        "C3 Broken Image Paths",
        "FCE4D6",
        ["patient_id", "image_path", "issue_detail"],
    )
    _sheet_null_cols(wb, c4)
    _sheet_fidelity(wb, c5)
    _sheet_dropped(wb, dropped_df)
    _sheet_missing(wb, missing_df)
    _sheet_unmapped(wb, col_coverage)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    result = "PASSED" if overall else "FAILED - review report"
    print(f"\n  Result : {result}")
    print(f"  Saved  : {output_path}\n")
    return overall


def _parse_args():
    p = argparse.ArgumentParser(
        description="Validate SMART-II patient data cleaning: XLSX vs cleaned CSV."
    )
    p.add_argument("--xlsx", help="Path to original XLSX")
    p.add_argument("--csv", help="Path to cleaned CSV")
    p.add_argument("--images", help="Base images directory")
    p.add_argument("--output", help="Path for validation report XLSX")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    config = load_config()

    basepath = config.get("SMART-II-DATAPATH", "smart.basepath")
    xlsx_path = (
        args.xlsx
        or f"{basepath}/{config.get('SMART-II-DATAPATH', 'smart.patient.metadata')}"
    )
    csv_path = args.csv or config.get(
        "SMART-II-DATAPATH", "smart.patient.clean.metadata.filename"
    )
    images_dir = args.images or basepath
    validate_report = args.output or config.get(
        "VALIDATE", "validate.smart.patient.report"
    )

    success = run(
        xlsx_path=xlsx_path,
        csv_path=csv_path,
        images_dir=images_dir,
        output_path=validate_report,
    )
    raise SystemExit(0 if success else 1)
